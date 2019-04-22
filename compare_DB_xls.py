# -*- coding: utf-8 -*-
"""
Created on Wed Nov  7 15:55:19 2018

classes to compare CMACS DB entries with the "golden source" stored in XLS format

@author: kyvolodk
"""

import pyodbc
import openpyxl as xl
#import collections

class CmacsDBQuery:
    def __init__(self, dbQuery, env):
        self.dbQuery = dbQuery
        self.env = env
    def run(self):       
        cmacs = pyodbc.connect('Driver={SQL Server};'
                               'Server='+self.env+';'
                               'Database=CMACS;'
                               'Trusted_Connection=yes;')
        cursor = cmacs.cursor()
        cursor.execute(self.dbQuery)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        cmacs.close()
        return [columns, rows];

class DbXlsCompare:
    def __init__(self, dbQuery, env, xlsFile):
        self.dbQuery = dbQuery
        self.env = env
        self.xlsFile = xlsFile
    
    def getDBHeaders(self):
        query = CmacsDBQuery(self.dbQuery, self.env)
        columns = list(query.run()[0])
        return columns;
    
    def getXLSHeaders(self):
        ws = xl.load_workbook(filename = self.xlsFile).get_sheet_by_name('Sheet1')
        columns = [cell.value for cell in list(ws[1])]
        return columns;
    
    def getDBResults(self):
        query = CmacsDBQuery(self.dbQuery, self.env)
        rawResults = query.run()[1]
        results = []
        for row in rawResults:
            resultRow = []
            for item in row:
                resultRow.append(item)
            resultRow = tuple(resultRow)
            results.append(resultRow)            
        return results;         
        
    def getXLSResults(self):
        ws = xl.load_workbook(filename = self.xlsFile).get_sheet_by_name('Sheet1')
        rawResults = list(ws.iter_rows(min_row = 2))
        results = []
        #resultRow = ()
        
        for row in rawResults:
            resultRow = []
            for cell in row:
                resultRow.append(cell.value)
            resultRow = tuple(resultRow)
            results.append(resultRow)
            #print(results)            
        return results;         
               
    def compare (self, resultsFile = None):
        dbResults = sorted(self.getDBResults())
        xlsResults = sorted(self.getXLSResults())
        #print(dbResults)
        #print(xlsResults)
        if dbResults == xlsResults:
            print('No discrepancies')
            return True
        else:
            db_notXls = list(set(dbResults) - set(xlsResults))
            xls_notDb = list(set(xlsResults) - set(dbResults))
            if resultsFile == None:
                print('In DB but not in XLS:')
                print(self.getDBHeaders())
                print(db_notXls)
                print('In XLS but not in DB:')
                print(self.getXLSHeaders())
                print(xls_notDb) 
            else:
                wb = xl.Workbook()
                wb.create_sheet('ComparisonResults')
                wb.remove(wb.get_sheet_by_name('Sheet'))
                ws = wb.get_sheet_by_name('ComparisonResults')
                ws['A1'] = 'In DB but not in XLS:'
                ws.append(self.getDBHeaders())
                for row in db_notXls:
                    ws.append(row)
                ws.append([])
                ws.append([])
                ws.append(['In XLS but not in DB:'])
                ws.append(self.getXLSHeaders())
                for row in xls_notDb:
                    ws.append(row)
                wb.save(resultsFile)
                wb.close()
                print('Discrepancies generated in '+resultsFile)
                
            return False        

        